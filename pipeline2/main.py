import os
import json
import glob
import threading
import time
import re
import tempfile
import asyncio
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
from matplotlib.figure import Figure
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from call_llm import call_deepseek, call_doubao, call_qwen, call_deepseek_v3, call_deepseek_json, \
                  async_call_deepseek, async_call_doubao, async_call_qwen, async_call_deepseek_v3
from prompts import *
from md_latex_renderer import MdLatexRenderer
# 定义提示词模板

# 配置matplotlib使用LaTeX渲染
matplotlib.rcParams.update({
    'text.usetex': True,
    'font.family': 'serif',
    'font.serif': ['Computer Modern Roman'],
    'mathtext.fontset': 'cm',
    'axes.unicode_minus': False
})

# 正则化提取result：result[key]
def get_key_value(text):
    result = {}
    
    # 匹配数组类型的键值对
    array_pattern = r'"([^"]+)"\s*:\s*\[(.*?)\]'
    array_matches = re.findall(array_pattern, text)
    for key, array_content in array_matches:
        # 提取数组中的每个元素
        array_items = re.findall(r'"([^"]+)"', array_content)
        result[key] = array_items
    
    # 匹配包含复杂内容的普通键值对（如包含LaTeX公式的文本）
    # 使用非贪婪匹配来处理包含引号和特殊字符的内容
    string_pattern = r'"([^"]+)"\s*:\s*"(.*?)(?:"|$)'
    string_matches = re.findall(string_pattern, text, re.DOTALL)
    for key, value in string_matches:
        if key not in result:  # 避免覆盖已匹配的数组
            result[key] = value
    
    return result

def get_ans(question_json):
    """
    获取三个模型的答案并更新问题JSON
    
    Args:
        question_json: 题目的JSON对象，会被直接修改
    """
    # Combine the 题干 and 提问 to form the complete question
    question = question_json["题干"] + " " + question_json["提问"]
    print(f"Processing question: {question}")
    
    # 使用协程并行调用三个模型
    model_times = {"deepseek": 0, "doubao": 0, "qwen": 0}
    
    async def call_models_async():
        # 创建三个模型的异步任务
        print("开始并行调用三个模型...")
        
        async def call_deepseek_async():
            print(f"Calling DeepSeek...")
            model_start = time.time()
            result = await async_call_deepseek(question)
            model_end = time.time()
            model_times["deepseek"] = model_end - model_start
            print(f"DeepSeek completed in {model_times['deepseek']:.2f} seconds")
            return result
            
        async def call_doubao_async():
            print(f"Calling Doubao...")
            model_start = time.time()
            result = await async_call_doubao(question)
            model_end = time.time()
            model_times["doubao"] = model_end - model_start
            print(f"Doubao completed in {model_times['doubao']:.2f} seconds")
            return result
            
        async def call_qwen_async():
            print(f"Calling Qwen...")
            model_start = time.time()
            result = await async_call_qwen(question)
            model_end = time.time()
            model_times["qwen"] = model_end - model_start
            print(f"Qwen completed in {model_times['qwen']:.2f} seconds")
            return result
        
        # 并行运行所有任务
        deepseek_result, doubao_result, qwen_result = await asyncio.gather(
            call_deepseek_async(), call_doubao_async(), call_qwen_async()
        )
        
        return deepseek_result, doubao_result, qwen_result
    
    # 使用asyncio.run运行异步函数
    try:
        # 设置事件循环策略以避免Windows上的警告
        if os.name == 'nt':
            asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
            
        # 运行异步函数
        deepseek_ans, doubao_ans, qwen_ans = asyncio.run(call_models_async())
        
        # Add the answers to the JSON
        question_json["deepseek_ans"] = deepseek_ans
        question_json["doubao_ans"] = doubao_ans
        question_json["qwen_ans"] = qwen_ans
        
        # Add timing information to the JSON
        question_json["model_times"] = model_times
    except Exception as e:
        print(f"调用模型时出错: {e}")
        # 如果协程方法失败，回退到线程方法
        print("回退到线程方法...")
        get_ans_with_threads(question_json)

def get_ans_with_threads(question_json):
    """
    使用线程获取三个模型的答案并更新问题JSON（作为备用方法）
    
    Args:
        question_json: 题目的JSON对象，会被直接修改
    """
    # Combine the 题干 and 提问 to form the complete question
    question = question_json["题干"] + " " + question_json["提问"]
    print(f"Processing question with threads: {question}")
    
    # Use threads to call the three models in parallel
    results = {"deepseek_ans": None, "doubao_ans": None, "qwen_ans": None}
    model_times = {"deepseek": 0, "doubao": 0, "qwen": 0}
    
    def call_deepseek_thread():
        print(f"Calling DeepSeek...")
        model_start = time.time()
        results["deepseek_ans"] = call_deepseek(question)
        model_end = time.time()
        model_times["deepseek"] = model_end - model_start
        print(f"DeepSeek completed in {model_times['deepseek']:.2f} seconds")
        
    def call_doubao_thread():
        print(f"Calling Doubao...")
        model_start = time.time()
        results["doubao_ans"] = call_doubao(question)
        model_end = time.time()
        model_times["doubao"] = model_end - model_start
        print(f"Doubao completed in {model_times['doubao']:.2f} seconds")
        
    def call_qwen_thread():
        print(f"Calling Qwen...")
        model_start = time.time()
        results["qwen_ans"] = call_qwen(question)
        model_end = time.time()
        model_times["qwen"] = model_end - model_start
        print(f"Qwen completed in {model_times['qwen']:.2f} seconds")
    
    # Create threads
    threads = [
        threading.Thread(target=call_deepseek_thread),
        threading.Thread(target=call_doubao_thread),
        threading.Thread(target=call_qwen_thread)
    ]
    
    # Start all threads
    for thread in threads:
        thread.start()
    
    # Wait for all threads to complete
    for thread in threads:
        thread.join()
    
    # Add the answers to the JSON
    question_json["deepseek_ans"] = results["deepseek_ans"]
    question_json["doubao_ans"] = results["doubao_ans"]
    question_json["qwen_ans"] = results["qwen_ans"]
    
    # Add timing information to the JSON
    question_json["model_times"] = model_times

def jude_ans(question_json):
    """
    判断三个模型的答案是否正确
    
    Args:
        question_json: 已经被get_ans处理过的题目JSON对象，会被直接修改
    
    Returns:
        一个字典，包含每个模型答案的评估结果（"正确"或"错误"）
    """
    # 获取问题、标准答案和三个模型的答案
    question = question_json["题干"] + " " + question_json["提问"]
    standard_answer = question_json["最终答案"]
    deepseek_ans = question_json["deepseek_ans"]
    doubao_ans = question_json["doubao_ans"]
    qwen_ans = question_json["qwen_ans"]
    
    # 创建评估结果字典
    results = {"deepseek": None, "doubao": None, "qwen": None}
    model_times = {"deepseek_v3": 0, "doubao_v3": 0, "qwen_v3": 0}
    full_responses = {"deepseek": "", "doubao": "", "qwen": ""}
    
    # 定义线程评估方法（备用方法）
    def eval_with_threads():
        """使用线程并行评估三个模型的答案（备用方法）"""
        def eval_deepseek_thread():
            print(f"Evaluating DeepSeek answer with thread...")
            model_start = time.time()
            prompt = create_eval_prompt("DeepSeek", deepseek_ans)
            response = call_deepseek_v3(prompt)
            full_responses["deepseek"] = response
            results["deepseek"] = extract_result(response)
            model_end = time.time()
            model_times["deepseek_v3"] = model_end - model_start
            print(f"DeepSeek evaluation completed in {model_times['deepseek_v3']:.2f} seconds")
            
        def eval_doubao_thread():
            print(f"Evaluating Doubao answer with thread...")
            model_start = time.time()
            prompt = create_eval_prompt("Doubao", doubao_ans)
            response = call_deepseek_v3(prompt)
            full_responses["doubao"] = response
            results["doubao"] = extract_result(response)
            model_end = time.time()
            model_times["doubao_v3"] = model_end - model_start
            print(f"Doubao evaluation completed in {model_times['doubao_v3']:.2f} seconds")
            
        def eval_qwen_thread():
            print(f"Evaluating Qwen answer with thread...")
            model_start = time.time()
            prompt = create_eval_prompt("Qwen", qwen_ans)
            response = call_deepseek_v3(prompt)
            full_responses["qwen"] = response
            results["qwen"] = extract_result(response)
            model_end = time.time()
            model_times["qwen_v3"] = model_end - model_start
            print(f"Qwen evaluation completed in {model_times['qwen_v3']:.2f} seconds")
        
        # 创建线程
        threads = [
            threading.Thread(target=eval_deepseek_thread),
            threading.Thread(target=eval_doubao_thread),
            threading.Thread(target=eval_qwen_thread)
        ]
        
        # 启动所有线程
        for thread in threads:
            thread.start()
        
        # 等待所有线程完成
        for thread in threads:
            thread.join()
    
    # 创建评估提示模板
    def create_eval_prompt(model_name, model_answer):
        return f"""
请评估以下问题的答案是否正确：

问题: {question}
标准答案: {standard_answer}
模型的答案: {model_answer}

请以"所以，模型的答案是正确/错误的"结尾。
"""
    
    # 从模型回复中提取评估结果
    def extract_result(response):
        # 保存完整回复
        
        # 获取检查范围，避免索引错误
        check_range = response[-min(50, len(response)):]
        
        # 查找最后一句话中是否包含"正确"或"错误"
        if "模型的答案是正确的" in check_range:
            return "正确"
        elif "模型的答案是错误的" in check_range:
            return "错误"
        else:
            # 如果无法确定，尝试查找整个回复
            if "正确" in response and "错误" not in response:
                return "正确"
            elif "错误" in response and "正确" not in response:
                return "错误"
            else:
                # 如果仍无法确定，返回未知
                return "未知"
    
    # 使用协程并行评估三个模型的答案
    async def eval_models_async():
        async def eval_deepseek_async():
            print(f"Evaluating DeepSeek answer...")
            model_start = time.time()
            prompt = create_eval_prompt("DeepSeek", deepseek_ans)
            response = await async_call_deepseek_v3(prompt)
            full_responses["deepseek"] = response
            result = extract_result(response)
            model_end = time.time()
            model_times["deepseek_v3"] = model_end - model_start
            print(f"DeepSeek evaluation completed in {model_times['deepseek_v3']:.2f} seconds")
            return result
        
        async def eval_doubao_async():
            print(f"Evaluating Doubao answer...")
            model_start = time.time()
            prompt = create_eval_prompt("Doubao", doubao_ans)
            response = await async_call_deepseek_v3(prompt)
            full_responses["doubao"] = response
            result = extract_result(response)
            model_end = time.time()
            model_times["doubao_v3"] = model_end - model_start
            print(f"Doubao evaluation completed in {model_times['doubao_v3']:.2f} seconds")
            return result
        
        async def eval_qwen_async():
            print(f"Evaluating Qwen answer...")
            model_start = time.time()
            prompt = create_eval_prompt("Qwen", qwen_ans)
            response = await async_call_deepseek_v3(prompt)
            full_responses["qwen"] = response
            result = extract_result(response)
            model_end = time.time()
            model_times["qwen_v3"] = model_end - model_start
            print(f"Qwen evaluation completed in {model_times['qwen_v3']:.2f} seconds")
            return result
        
        # 并行运行所有评估任务
        deepseek_result, doubao_result, qwen_result = await asyncio.gather(
            eval_deepseek_async(), eval_doubao_async(), eval_qwen_async()
        )
        
        return deepseek_result, doubao_result, qwen_result
    
    # 定义线程评估方法（备用方法）
    def eval_with_threads():
        """使用线程并行评估三个模型的答案（备用方法）"""
        def eval_deepseek_thread():
            print(f"Evaluating DeepSeek answer with thread...")
            model_start = time.time()
            prompt = create_eval_prompt("DeepSeek", deepseek_ans)
            response = call_deepseek_v3(prompt)
            full_responses["deepseek"] = response
            results["deepseek"] = extract_result(response)
            model_end = time.time()
            model_times["deepseek_v3"] = model_end - model_start
            print(f"DeepSeek evaluation completed in {model_times['deepseek_v3']:.2f} seconds")
            
        def eval_doubao_thread():
            print(f"Evaluating Doubao answer with thread...")
            model_start = time.time()
            prompt = create_eval_prompt("Doubao", doubao_ans)
            response = call_deepseek_v3(prompt)
            full_responses["doubao"] = response
            results["doubao"] = extract_result(response)
            model_end = time.time()
            model_times["doubao_v3"] = model_end - model_start
            print(f"Doubao evaluation completed in {model_times['doubao_v3']:.2f} seconds")
            
        def eval_qwen_thread():
            print(f"Evaluating Qwen answer with thread...")
            model_start = time.time()
            prompt = create_eval_prompt("Qwen", qwen_ans)
            response = call_deepseek_v3(prompt)
            full_responses["qwen"] = response
            results["qwen"] = extract_result(response)
            model_end = time.time()
            model_times["qwen_v3"] = model_end - model_start
            print(f"Qwen evaluation completed in {model_times['qwen_v3']:.2f} seconds")
        
        # 创建线程
        threads = [
            threading.Thread(target=eval_deepseek_thread),
            threading.Thread(target=eval_doubao_thread),
            threading.Thread(target=eval_qwen_thread)
        ]
        
        # 启动所有线程
        for thread in threads:
            thread.start()
        
        # 等待所有线程完成
        for thread in threads:
            thread.join()
    
    # 使用asyncio.run运行异步函数
    try:
        # 设置事件循环策略以避免Windows上的警告
        if os.name == 'nt':
            asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
        
        # 运行异步函数
        deepseek_result, doubao_result, qwen_result = asyncio.run(eval_models_async())
        results["deepseek"] = deepseek_result
        results["doubao"] = doubao_result
        results["qwen"] = qwen_result
    except Exception as e:
        print(f"异步评估出错: {e}")
        # 如果协程方法失败，回退到线程方法
        print("回退到线程方法...")
        eval_with_threads()
    
    # 将评估结果添加到JSON
    question_json["evaluation"] = results
    question_json["evaluation_times"] = model_times
    question_json["evaluation_full_responses"] = full_responses
    
    return results

def extract_grade_and_subject(question, data):
    """提取适合年级和子学科"""
    flag = True
    count = 0
    while flag:
        try:
            print(f"提取适合年级和子学科...第{count+1}次尝试")
            response = call_deepseek_v3(prompt_question + question)
            response = get_key_value(response)
            data["适合年级"] = response["适合年级"]
            data["子学科"] = response["子学科"]
            print(f"适合年级: {data['适合年级']}, 子学科: {data['子学科']}")
            flag = False
            return True
        except Exception as e:
            print(f"提取适合年级和子学科失败: {e}")
            count += 1
            if count > 5:
                print("提取适合年级和子学科失败次数过多，放弃处理此题")
                return False
    return True

def extract_knowledge_and_analysis(question, standard_answer, data):
    """提取考察知识点和分析过程"""
    flag = True
    count = 0
    while flag:
        try:
            print(f"提取考察知识点和分析过程...第{count+1}次尝试")
            response = call_deepseek_v3(prompt_answer + question + "\n参考答案:" + standard_answer)
            response = get_key_value(response)
            data["考察知识点"] = response["考察知识点"]
            data["分析过程"] = response["分析过程"]
            print(f"考察知识点: {data['考察知识点']}")
            flag = False
            return True
        except Exception as e:
            print(f"提取考察知识点和分析过程失败: {e}")
            count += 1
            if count > 5:
                print("提取考察知识点和分析过程失败次数过多，放弃处理此题")
                return False
    return True

def extract_error_method_and_points(question, wrong_ans, standard_answer, data):
    """提取错误解题方法和易错点"""
    if not wrong_ans:
        return True
        
    flag = True
    count = 0
    while flag:
        try:
            print(f"提取错误解题方法和易错点...第{count+1}次尝试")
            response = call_deepseek_v3(prompt_wrong + question + "\n错误答案:" + wrong_ans + "\n正确答案:" + standard_answer)
            response = get_key_value(response)
            data["错误解题方法"] = response["错误解题方法"]
            data["易错点"] = response["易错点"]
            print(f"易错点: {data['易错点']}")
            flag = False
            return True
        except Exception as e:
            print(f"提取错误解题方法和易错点失败: {e}")
            count += 1
            if count > 5:
                print("提取错误解题方法和易错点失败次数过多，继续处理")
                break
    return True

def process_evaluation_results(question_json, data):
    """
    处理三个模型的评估结果
    
    Args:
        question_json: 包含评估结果的JSON对象
        data: 要更新的数据字典
    
    Returns:
        如果三个模型都正确，返回None；否则返回更新后的data
    """
    if "evaluation" not in question_json:
        return data
    
    # 收集正确的模型答案，用于设置解题过程
    correct_answers = []
    
    # 检查各模型的正确性
    deepseek_correct = question_json["evaluation"]["deepseek"] == "正确"
    doubao_correct = question_json["evaluation"]["doubao"] == "正确"
    qwen_correct = question_json["evaluation"]["qwen"] == "正确"
    
    # 如果模型正确，加入正确答案列表
    if deepseek_correct:
        correct_answers.append(("deepseek", question_json["deepseek_ans"]))
    if doubao_correct:
        correct_answers.append(("doubao", question_json["doubao_ans"]))
    if qwen_correct:
        correct_answers.append(("qwen", question_json["qwen_ans"]))
    
    # 如果三个模型都正确
    if deepseek_correct and doubao_correct and qwen_correct:
        print("三模型都正确")
        # 即使三个模型都正确，也选一个作为解题过程后返回
        if correct_answers:
            # 优先选择DeepSeek答案作为解题过程
            for model_name, answer in correct_answers:
                if model_name == "deepseek":
                    data["解题过程"] = answer
                    break
            # 如果没找到DeepSeek，就用第一个正确答案
            if not data["解题过程"]:
                data["解题过程"] = correct_answers[0][1]
        return None
    
    # 设置解题过程，优先从正确的模型中选择，按DeepSeek > Doubao > Qwen的优先级
    if correct_answers:
        # 查找是否有DeepSeek正确答案
        deepseek_answer = next((answer for model, answer in correct_answers if model == "deepseek"), None)
        if deepseek_answer:
            data["解题过程"] = deepseek_answer
        else:
            # 否则使用第一个正确答案
            data["解题过程"] = correct_answers[0][1]
    else:
        # 如果没有正确答案，使用标准答案
        data["解题过程"] = data["最终答案"]
    
    # 检查Qwen
    if not qwen_correct:
        data["三模型打分"] += 1
        data["错误解题方法"] = question_json["qwen_ans"]
        data["错误解法模型"] = "千问"
        data["错误模型"] = "千问"

    # 检查DeepSeek
    if not deepseek_correct:
        data["三模型打分"] += 1
        if not data["错误解题方法"]:
            data["错误解题方法"] = question_json["deepseek_ans"]
        if not data["错误解法模型"]:
            data["错误解法模型"] = "ds"
        if not data["错误模型"]:
            data["错误模型"] = "ds"
        else:
            data["错误模型"] += ",ds"
    
    # 检查Doubao
    if not doubao_correct:
        data["三模型打分"] += 1
        if not data["错误解题方法"]:
            data["错误解题方法"] = question_json["doubao_ans"]
        if not data["错误解法模型"]:
            data["错误解法模型"] = "豆包"
        if not data["错误模型"]:
            data["错误模型"] = "豆包"
        else:
            data["错误模型"] += ",豆包"
            
    return data

def get_png(question_id, question, deepseek_ans, doubao_ans, qwen_ans, deepseek_time=None, doubao_time=None, qwen_time=None, output_dir="screenshots", evaluation=None):
    """
    将模型的输出以Markdown格式保存为PNG文件，使用MdLatexRenderer渲染LaTeX公式
    仅为回答错误的模型生成PNG图片
    
    Args:
        question_id: 问题ID
        question: 问题内容
        deepseek_ans: DeepSeek模型的回答
        doubao_ans: Doubao模型的回答
        qwen_ans: Qwen模型的回答
        deepseek_time: DeepSeek模型处理时间
        doubao_time: Doubao模型处理时间
        qwen_time: Qwen模型处理时间
        output_dir: 输出目录
        evaluation: 评估结果字典，包含各模型的正确性判断
    
    Returns:
        包含错误模型PNG文件路径的字典
    """
    # 导入MdLatexRenderer
    try:
        renderer = MdLatexRenderer(output_dir=output_dir)
    except Exception as e:
        print(f"创建渲染器实例时出错: {str(e)}，将使用备用方法")
        return get_png_matplotlib(question_id, question, deepseek_ans, doubao_ans, qwen_ans, 
                                 deepseek_time, doubao_time, qwen_time, output_dir, evaluation)
    
    # 确保输出目录存在
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"创建输出目录: {output_dir}")
    
    # 创建每个模型的子目录
    deepseek_dir = os.path.join(output_dir, "deepseek")
    doubao_dir = os.path.join(output_dir, "doubao")
    qwen_dir = os.path.join(output_dir, "qwen")
    
    for directory in [deepseek_dir, doubao_dir, qwen_dir]:
        if not os.path.exists(directory):
            os.makedirs(directory)
    
    # 文件路径
    deepseek_txt = os.path.join(deepseek_dir, f"deepseek_text_{question_id}.md")
    doubao_txt = os.path.join(doubao_dir, f"doubao_text_{question_id}.md")
    qwen_txt = os.path.join(qwen_dir, f"qwen_text_{question_id}.md")
    
    deepseek_png = os.path.join(deepseek_dir, f"deepseek_screenshot_{question_id}.png")
    doubao_png = os.path.join(doubao_dir, f"doubao_screenshot_{question_id}.png")
    qwen_png = os.path.join(qwen_dir, f"qwen_screenshot_{question_id}.png")
    
    # 保存文本内容到Markdown文件
    def save_text_to_file(model_name, model_answer, filepath, processing_time=None):
        content = f"# {model_name}\n\n"
        content += f"## 题目\n{question}\n\n"
        content += f"## 回答\n{model_answer}\n\n"
        
        # 添加处理时间
        if processing_time is not None:
            content += f"## 处理时间\n{processing_time:.2f}秒\n"
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(content)
        return content
    
    # 根据模型的正确性判断是否需要生成图片
    deepseek_correct = True
    doubao_correct = True
    qwen_correct = True
    
    if evaluation:
        deepseek_correct = evaluation.get("deepseek") == "正确"
        doubao_correct = evaluation.get("doubao") == "正确"
        qwen_correct = evaluation.get("qwen") == "正确"
    
    # 默认使用文本文件路径
    paths = {
        "deepseek": None,
        "doubao": None,
        "qwen": None
    }
    
    try:
        # 只为错误的模型生成PNG图片
        if not deepseek_correct:
            deepseek_content = save_text_to_file("DeepSeek-R1", deepseek_ans, deepseek_txt, deepseek_time)
            # 确保临时目录存在
            temp_dir = os.path.join(tempfile.gettempdir(), "deepseek")
            os.makedirs(temp_dir, exist_ok=True)
            renderer.render_to_image(
                deepseek_content, 
                output_filename=os.path.join("deepseek", os.path.basename(deepseek_png)), 
                width=800,
                llm_name="DeepSeek-R1"
            )
            paths["deepseek"] = deepseek_png
            print(f"已生成DeepSeek错误回答的PNG截图")
            
        if not doubao_correct:
            doubao_content = save_text_to_file("Doubao-1-5-thinking-pro-250415", doubao_ans, doubao_txt, doubao_time)
            # 确保临时目录存在
            temp_dir = os.path.join(tempfile.gettempdir(), "doubao")
            os.makedirs(temp_dir, exist_ok=True)
            renderer.render_to_image(
                doubao_content, 
                output_filename=os.path.join("doubao", os.path.basename(doubao_png)), 
                width=800,
                llm_name="Doubao-1-5-thinking-pro-250415"
            )
            paths["doubao"] = doubao_png
            print(f"已生成豆包错误回答的PNG截图")
            
        if not qwen_correct:
            qwen_content = save_text_to_file("QwQ-32B", qwen_ans, qwen_txt, qwen_time)
            # 确保临时目录存在
            temp_dir = os.path.join(tempfile.gettempdir(), "qwen")
            os.makedirs(temp_dir, exist_ok=True)
            renderer.render_to_image(
                qwen_content, 
                output_filename=os.path.join("qwen", os.path.basename(qwen_png)), 
                width=800,
                llm_name="QwQ-32B"
            )
            paths["qwen"] = qwen_png
            print(f"已生成千问错误回答的PNG截图")
        
    except Exception as e:
        print(f"MdLatexRenderer渲染PNG时出错: {str(e)}")
    
    print(f"已保存错误模型的输出至 {output_dir} 目录")
    return paths

# 保留原始的matplotlib渲染方法作为备用
def get_png_matplotlib(question_id, question, deepseek_ans, doubao_ans, qwen_ans, deepseek_time=None, doubao_time=None, qwen_time=None, output_dir="screenshots", evaluation=None):
    """
    将模型的输出以Markdown格式保存为PNG文件（使用matplotlib作为备用渲染方法）
    仅为回答错误的模型生成PNG图片
    
    Args:
        question_id: 问题ID
        question: 问题内容
        deepseek_ans: DeepSeek模型的回答
        doubao_ans: Doubao模型的回答
        qwen_ans: Qwen模型的回答
        deepseek_time: DeepSeek模型处理时间
        doubao_time: Doubao模型处理时间
        qwen_time: Qwen模型处理时间
        output_dir: 输出目录
        evaluation: 评估结果字典，包含各模型的正确性判断
    
    Returns:
        包含错误模型PNG文件路径的字典
    """
    # 确保输出目录存在
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"创建输出目录: {output_dir}")
    
    # 创建每个模型的子目录
    deepseek_dir = os.path.join(output_dir, "deepseek")
    doubao_dir = os.path.join(output_dir, "doubao")
    qwen_dir = os.path.join(output_dir, "qwen")
    
    for directory in [deepseek_dir, doubao_dir, qwen_dir]:
        if not os.path.exists(directory):
            os.makedirs(directory)
    
    # 获取当前时间戳
    current_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    
    # 根据模型的正确性判断是否需要生成图片
    deepseek_correct = True
    doubao_correct = True
    qwen_correct = True
    
    if evaluation:
        deepseek_correct = evaluation.get("deepseek") == "正确"
        doubao_correct = evaluation.get("doubao") == "正确"
        qwen_correct = evaluation.get("qwen") == "正确"
    
    # 默认为空
    paths = {
        "deepseek": None,
        "doubao": None,
        "qwen": None
    }
    
    # 保存为文本文件（作为备份）
    def save_text_to_file(model_name, model_answer, filepath):
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(f"# {model_name}\n\n")
            f.write(f"## 题目\n{question}\n\n")
            f.write(f"## 回答\n{model_answer}\n\n")
            f.write(f"## 时间\n{current_time}\n")
            
            # 如果有处理时间，也添加进去
            if model_name == "DeepSeek-R1" and deepseek_time is not None:
                f.write(f"\n## 处理时间\n{deepseek_time:.2f}秒\n")
            elif model_name == "Doubao-1-5-thinking-pro-250415" and doubao_time is not None:
                f.write(f"\n## 处理时间\n{doubao_time:.2f}秒\n")
            elif model_name == "QwQ-32B" and qwen_time is not None:
                f.write(f"\n## 处理时间\n{qwen_time:.2f}秒\n")
    
    # 文本文件路径
    deepseek_txt = os.path.join(deepseek_dir, f"deepseek_text_{question_id}.md")
    doubao_txt = os.path.join(doubao_dir, f"doubao_text_{question_id}.md")
    qwen_txt = os.path.join(qwen_dir, f"qwen_text_{question_id}.md")
    
    # PNG文件路径
    deepseek_png = os.path.join(deepseek_dir, f"deepseek_screenshot_{question_id}.png")
    doubao_png = os.path.join(doubao_dir, f"doubao_screenshot_{question_id}.png")
    qwen_png = os.path.join(qwen_dir, f"qwen_screenshot_{question_id}.png")
    
    # 将文本直接渲染为图像
    def save_as_image(title, content, time_str, processing_time, output_path):
        try:
            # 创建图像
            dpi = 100
            fig_width = 10
            fig_height = 15  # 增加高度以容纳更多内容
            
            # 将内容分成多行
            lines = content.split('\n')
            # 估算需要的高度
            estimated_height = len(lines) * 0.3 + 6  # 给标题和边距预留额外空间
            fig_height = max(fig_height, estimated_height)
            
            fig = plt.figure(figsize=(fig_width, fig_height), dpi=dpi)
            ax = fig.add_subplot(111)
            ax.axis('off')
            
            # 设置背景色为白色
            ax.set_facecolor('white')
            fig.patch.set_facecolor('white')
            
            # 添加标题和内容
            y_pos = 0.98
            ax.text(0.05, y_pos, title, fontsize=18, weight='bold', transform=ax.transAxes)
            y_pos -= 0.04
            
            # 添加题目标题
            y_pos -= 0.04
            ax.text(0.05, y_pos, "题目:", fontsize=14, weight='bold', transform=ax.transAxes)
            y_pos -= 0.04
            
            # 添加题目内容
            question_lines = question.split('\n')
            max_question_length = 80
            question_content = []
            for line in question_lines:
                while len(line) > max_question_length:
                    question_content.append(line[:max_question_length])
                    line = line[max_question_length:]
                question_content.append(line)
            
            # 添加题目文本
            chunk = '\n'.join(question_content)
            ax.text(0.05, y_pos, chunk, fontsize=10, transform=ax.transAxes,
                  verticalalignment='top', wrap=True)
            y_pos -= 0.10  # 根据文本块大小调整位置
            
            # 添加模型回答标题
            y_pos -= 0.04
            ax.text(0.05, y_pos, "回答:", fontsize=14, weight='bold', transform=ax.transAxes)
            y_pos -= 0.04
            
            # 添加内容（将过长的内容分多次添加，避免超出边界）
            max_line_length = 80  # 每行最大字符数
            content_lines = []
            
            for line in lines:
                # 处理过长的行
                while len(line) > max_line_length:
                    content_lines.append(line[:max_line_length])
                    line = line[max_line_length:]
                content_lines.append(line)
            
            # 每次添加一块文本
            chunk_size = 30  # 每次添加的行数
            for i in range(0, len(content_lines), chunk_size):
                chunk = '\n'.join(content_lines[i:i + chunk_size])
                ax.text(0.05, y_pos, chunk, fontsize=10, transform=ax.transAxes,
                      verticalalignment='top', wrap=True)
                y_pos -= 0.25  # 根据文本块大小调整位置
            
            # 添加时间和处理时间
            y_pos = max(0.05, y_pos)  # 确保不超出底部边界
            ax.text(0.05, y_pos, f"时间: {time_str}", fontsize=10, transform=ax.transAxes)
            
            if processing_time is not None:
                y_pos -= 0.04
                ax.text(0.05, y_pos, f"处理时间: {processing_time:.2f}秒", fontsize=10, transform=ax.transAxes)
            
            # 保存图像
            plt.tight_layout()
            fig.savefig(output_path, format='png', bbox_inches='tight', pad_inches=0.5)
            plt.close(fig)
            return True
        except Exception as e:
            print(f"生成图像时出错: {str(e)}")
            return False
    
    # 只为错误的模型生成PNG
    try:
        # DeepSeek
        if not deepseek_correct:
            save_text_to_file("DeepSeek-R1", deepseek_ans, deepseek_txt)
            if save_as_image("DeepSeek-R1", deepseek_ans, current_time, deepseek_time, deepseek_png):
                paths["deepseek"] = deepseek_png
                print(f"已生成DeepSeek错误回答的PNG截图")
            else:
                paths["deepseek"] = deepseek_txt
                
        # 豆包
        if not doubao_correct:
            save_text_to_file("Doubao-1-5-thinking-pro-250415", doubao_ans, doubao_txt)
            if save_as_image("Doubao-1-5-thinking-pro-250415", doubao_ans, current_time, doubao_time, doubao_png):
                paths["doubao"] = doubao_png
                print(f"已生成豆包错误回答的PNG截图")
            else:
                paths["doubao"] = doubao_txt
                
        # 千问
        if not qwen_correct:
            save_text_to_file("QwQ-32B", qwen_ans, qwen_txt)
            if save_as_image("QwQ-32B", qwen_ans, current_time, qwen_time, qwen_png):
                paths["qwen"] = qwen_png
                print(f"已生成千问错误回答的PNG截图")
            else:
                paths["qwen"] = qwen_txt
                
    except Exception as e:
        print(f"PNG生成过程中出错: {str(e)}")
    
    print(f"已保存错误模型的输出至 {output_dir} 目录")
    return paths

def extract_timu_data(question_json):
    """
    从评估结果中提取统计数据和额外分析数据
    
    Args:
        question_json: 处理完成的JSON对象，包含模型答案和评估结果
    
    Returns:
        包含统计和分析数据的字典
    """
    question = question_json["题干"] + " " + question_json["提问"]
    standard_answer = question_json["最终答案"]
    
    # 初始化数据
    data = {
        'id': question_json.get("序号", ""),
        "问题条件": question_json["题干"],
        "具体问题": question_json["提问"],
        "问题数目": 1,
        "适合年级": None,
        "题目类型": "计算题",
        "题目学科": "数学",
        "子学科": None,
        "领域类型": "自然科学",
        "是否包含图片": "否",
        "考察知识点": None,
        "易错点": None,
        "思考过程/分析": None,
        "解题过程": None,
        "最终答案": question_json["最终答案"],
        "错误解题方法": None,
        "错误解法模型": None,  # 错误解法模型
        "错误模型": "",       # 错误模型
        "三模型打分": 0,      # 错误答案数量
        "deepseek": None,
        "千问": None,
        "豆包": None,
        "题目来源": None,
        "deepseek_ans": question_json["deepseek_ans"],
        "doubao_ans": question_json["doubao_ans"],
        "qwen_ans": question_json["qwen_ans"],
    }
    
    # 处理评估结果
    data = process_evaluation_results(question_json, data)
    if data is None:
        return None
    
    # 生成模型回答的PNG截图
    question_id = question_json.get("序号", "unknown")
    model_times = question_json.get("model_times", {})
    png_paths = get_png(
        question_id=question_id,
        question=question,
        deepseek_ans=question_json["deepseek_ans"],
        doubao_ans=question_json["doubao_ans"],
        qwen_ans=question_json["qwen_ans"],
        deepseek_time=model_times.get("deepseek"),
        doubao_time=model_times.get("doubao"),
        qwen_time=model_times.get("qwen"),
        output_dir="screenshots",
        evaluation=question_json["evaluation"] if "evaluation" in question_json else None
    )
    
    # 更新PNG路径
    data["deepseek"] = png_paths["deepseek"]
    data["千问"] = png_paths["qwen"]
    data["豆包"] = png_paths["doubao"]
    
    # 准备错误答案
    wrong_ans = data["错误解题方法"]
    
    # 创建三个线程并行提取数据
    print("开始并行提取额外数据...")
    extraction_results = {"grade_subject": False, "knowledge_analysis": False, "error_points": False}
    
    # 定义线程函数
    def run_grade_subject_thread():
        extraction_results["grade_subject"] = extract_grade_and_subject(question, data)
        
    def run_knowledge_analysis_thread():
        extraction_results["knowledge_analysis"] = extract_knowledge_and_analysis(question, standard_answer, data)
        
    def run_error_points_thread():
        extraction_results["error_points"] = extract_error_method_and_points(question, wrong_ans, standard_answer, data)
    
    # 创建线程
    threads = [
        threading.Thread(target=run_grade_subject_thread),
        threading.Thread(target=run_knowledge_analysis_thread),
        threading.Thread(target=run_error_points_thread)
    ]
    
    # 启动所有线程
    for thread in threads:
        thread.start()
    
    # 等待所有线程完成
    for thread in threads:
        thread.join()
    
    # 检查关键提取是否成功
    if not extraction_results["grade_subject"] or not extraction_results["knowledge_analysis"]:
        print("关键数据提取失败，放弃处理此题")
        return None
    
    # 更新data中的额外数据（这一步现在其实不需要了，因为线程直接修改了data字典）
    data.update({
        "适合年级": data["适合年级"],
        "子学科": data["子学科"],
        "考察知识点": data["考察知识点"],
        "思考过程/分析": data["分析过程"],
        "易错点": data["易错点"]
    })
    
    # 更新到question_json
    question_json["stats"] = data
    question_json["data"] = data
    
    return data

def save_checkpoint(current_index, data_list, checkpoint_file="last_process_pipe2.npy"):
    """保存处理进度检查点"""
    try:
        np.save(checkpoint_file, np.array([current_index, data_list], dtype=object))
        print(f"检查点已保存: 当前索引={current_index}, 数据数量={len(data_list)}")
    except Exception as e:
        print(f"保存检查点出错: {e}")

def load_checkpoint(checkpoint_file="last_process_pipe2.npy"):
    """加载处理进度检查点"""
    try:
        if os.path.exists(checkpoint_file):
            checkpoint = np.load(checkpoint_file, allow_pickle=True)
            current_index = checkpoint[0]
            data_list = checkpoint[1].tolist() if isinstance(checkpoint[1], np.ndarray) else checkpoint[1]
            print(f"已加载检查点: 当前索引={current_index}, 数据数量={len(data_list)}")
            return current_index, data_list
        else:
            print("没有找到检查点文件，将从头开始处理")
            return 0, []
    except Exception as e:
        print(f"加载检查点出错: {e}")
        return 0, []

def save_to_excel(data_list, file_path):
    """保存数据到Excel文件"""
    try:
        # 转换为DataFrame
        df = pd.DataFrame(data_list)
        
        # 预处理数据，确保所有项都可以被Excel接受
        def safe_excel_value(value):
            if isinstance(value, list):
                return ', '.join(str(item) for item in value)
            elif value is None:
                return ""
            else:
                return str(value)
        
        # 应用安全转换，避免非法字符和列表类型
        for column in df.columns:
            df[column] = df[column].apply(safe_excel_value)
        
        # 清理非法字符
        illegal_chars_re = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")
        df = df.applymap(lambda x: illegal_chars_re.sub("", x) if isinstance(x, str) else x)
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        
        # 写入表头
        for col_idx, column in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=column)
        
        # 写入数据
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                
                # 检查是否是图片列
                column_name = df.columns[col_idx-1]
                if column_name in ['deepseek', '千问', '豆包'] and value and isinstance(value, str) and value.endswith('.png'):
                    # 直接使用相对路径，因为图片路径已经是相对于工作目录的
                    if os.path.exists(value):
                        try:
                            img = Image(value)
                            # 设置更大的图片尺寸
                            img.width = 400
                            img.height = 400
                            # 调整单元格大小以适应图片
                            ws.row_dimensions[row_idx].height = 300
                            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 50
                            ws.add_image(img, cell.coordinate)
                            cell.value = ""  # 清除单元格文本值
                            print(f"成功添加图片: {value} 到单元格 {cell.coordinate}")
                        except Exception as e:
                            print(f"添加图片到 {cell.coordinate} 时出错: {str(e)}")
                            print(f"图片路径: {value}")
                    else:
                        print(f"图片文件不存在: {value}")
        
        # 保存工作簿
        wb.save(file_path)
        print(f"Excel文件已保存: {file_path}")
    except Exception as e:
        print(f"保存Excel文件时出错: {str(e)}")
        # 尝试保存为简化版本，不包含图片
        try:
            # 创建备份文件名
            backup_path = file_path.replace('.xlsx', '_backup.xlsx')
            # 直接将DataFrame保存为Excel，跳过图片和复杂格式
            df.to_excel(backup_path, index=False)
            print(f"已保存备份Excel文件（不含图片）: {backup_path}")
        except Exception as e2:
            print(f"保存备份Excel文件也失败: {str(e2)}")

def main(start_index=0, checkpoint_file="last_process_pipe2.npy", output_dir="三模型表"):
    """
    处理timu目录下的所有JSON文件
    
    Args:
        start_index: 开始处理的文件索引
        checkpoint_file: 检查点文件路径
        output_dir: 输出目录
    """
    # 确保输出目录存在
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"创建输出目录: {output_dir}")
    
    # 加载检查点
    current_index, all_results = load_checkpoint(checkpoint_file)
    
    # Get all JSON files in the timu directory
    timu_files = glob.glob("timu/*.json")
    
    # 按数字顺序排序文件
    def extract_number(filename):
        # 从文件名中提取题号数字
        match = re.search(r'(\d+)_', os.path.basename(filename))
        if match:
            return int(match.group(1))
        return 0  # 如果无法提取数字，返回0
    
    timu_files.sort(key=extract_number)    
    total_start_time = time.time()
    
    # 从检查点继续处理
    for i, file_path in enumerate(timu_files[current_index:], start=current_index):
        file_start_time = time.time()
        print(f"Processing {i+1}/{len(timu_files)}: {file_path}...")
        
        # Load the JSON file
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Get answers from three models
        get_ans(data)
        
        # 立即保存包含三模型答案的JSON
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        print(f"Saved model answers to {file_path}")
        
        # Evaluate the answers
        evaluation = jude_ans(data)
        
        # 提取题目数据和额外分析数据
        timu_data = extract_timu_data(data)
        if timu_data is None:
            continue
        all_results.append(timu_data)
        
        # 保存包含评估结果和额外数据的JSON
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        
        file_end_time = time.time()
        file_duration = file_end_time - file_start_time
        print(f"Completed {file_path} in {file_duration:.2f} seconds")
        print(f"Model times: DeepSeek={data['model_times']['deepseek']:.2f}s, Doubao={data['model_times']['doubao']:.2f}s, Qwen={data['model_times']['qwen']:.2f}s")
        print(f"Evaluation: DeepSeek={evaluation['deepseek']}, Doubao={evaluation['doubao']}, Qwen={evaluation['qwen']}")
        print(f"Wrong answers: {timu_data['三模型打分']}/3, Wrong models: {timu_data['错误模型'] if timu_data['错误模型'] else '无'}")
        print("-" * 50)
        
        # 保存检查点（每处理一个文件后）
        save_checkpoint(i + 1, all_results, checkpoint_file)
        
        # 每5个文件保存一次Excel
        if (i + 1) % 5 == 0:
            now_time = time.strftime("%Y-%m-%d_%H%M%S", time.localtime())
            excel_path = os.path.join(output_dir, f"国内三模型_{now_time}.xlsx")
            save_to_excel(all_results, excel_path)
            print(f"已保存Excel文件: {excel_path}")
        
        # 测试模式：只处理一个文件后退出循环
        if "TEST_MODE" in os.environ:
            print("测试模式：只处理了一个文件")
            break
    
    # 保存总结果
    with open("timu_results.json", 'w', encoding='utf-8') as f:
        json.dump(all_results, f, ensure_ascii=False, indent=4)
    print(f"All results saved to timu_results.json")
    
    # 最终保存Excel文件
    now_time = time.strftime("%Y-%m-%d_%H%M%S", time.localtime())
    excel_path = os.path.join(output_dir, f"国内三模型_{now_time}.xlsx")
    save_to_excel(all_results, excel_path)
    print(f"已保存最终Excel文件: {excel_path}")
    
    # 计算总体统计信息
    total_questions = len(all_results)
    total_wrong = sum(item["三模型打分"] for item in all_results)
    models_accuracy = {
        "deepseek": sum(1 for item in all_results if item.get("deepseek_correct", False)) / total_questions,
        "doubao": sum(1 for item in all_results if item.get("doubao_correct", False)) / total_questions,
        "qwen": sum(1 for item in all_results if item.get("qwen_correct", False)) / total_questions
    }
    
    print("\n===== 总体统计 =====")
    print(f"总题目数: {total_questions}")
    print(f"总错误数: {total_wrong}")
    print(f"模型准确率: DeepSeek={models_accuracy['deepseek']:.2%}, Doubao={models_accuracy['doubao']:.2%}, Qwen={models_accuracy['qwen']:.2%}")
    
    total_end_time = time.time()
    total_duration = total_end_time - total_start_time
    print(f"All files processed in {total_duration:.2f} seconds")
    print(f"Average time per file: {total_duration / len(timu_files):.2f} seconds")
    print("所有问题处理完毕")

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="处理timu目录下的JSON文件")
    parser.add_argument("--start", type=int, default=0, help="开始处理的文件索引")
    parser.add_argument("--checkpoint", type=str, default="last_process_pipe2.npy", help="检查点文件路径")
    parser.add_argument("--output", type=str, default="三模型表", help="输出目录")
    parser.add_argument("--test", action="store_true", help="测试模式，只处理一个文件")
    
    args = parser.parse_args()
    
    if args.test:
        os.environ["TEST_MODE"] = "1"
    
    main(args.start, args.checkpoint, args.output)
